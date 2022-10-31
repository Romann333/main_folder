import re
import time
import requests
from bs4 import BeautifulSoup
from selenium import webdriver
from selenium.webdriver.common.by import By
import json
import csv
from openpyxl import load_workbook, Workbook
import random
import  os
import sqlite3
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support.expected_conditions import visibility_of_element_located
from datetime import datetime

NO_INFO_STATUS = 'No information'

DATA_FOLDER = 'Parsing/Venchur_fonds/'

PATH_TO_WEBDRIVER = "/home/roman/python_course/Selenium_Python/Chrome_ driver/chromedriver"


def get_all_links():

    headers = {"accept": "text/css,*/*;q=0.1", 'user-agent' : 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36'
    }


    for page in range(0, 10, 10): 
        fund_links = {}   
        gfs = requests.get(f'https://project-valentine-api.herokuapp.com/investors?page%5Blimit%5D=10&page%5Boffset%5D={page}',
        headers=headers
        )
    
        s = gfs.json()     
        
        for i in s['data']:
            temp = 'https://connect.visible.vc/investors/' + i['attributes']['slug']
            name = i['attributes']['name']
            fund_links[name] = temp

        if not os.path.exists(f"{DATA_FOLDER}data"):
            os.mkdir(f"{DATA_FOLDER}data")

        with open(f"{DATA_FOLDER}data/all_links - {page}.json", "w") as file:
            json.dump(fund_links, file, indent=4, ensure_ascii=False)

        time.sleep(random.randint(1, 2))

        print(f'Progress ... {page}')

def create_result_headers():

    HEADERS_RESULT_TABLE = (
        'Name of investor',
        'Site',
        'Info card',
        'Stage',
        'Check size',
        'Focus',
        'Investment geography',
        'Manager name',
        'Role',
        'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact',
        'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact', 'Manager name', 'Role', 'Contact'          
    )

    with open(f"{DATA_FOLDER}result.csv", 'w', encoding='utf-8') as file:  # Добавление заголовков в таблицу csv
        writer = csv.writer(file)
        writer.writerow(HEADERS_RESULT_TABLE)

    wb = Workbook()                                                         # Добавление заголовков в таблицу excel
    ws = wb.active
    ws.append(HEADERS_RESULT_TABLE) 
    wb.save(f"{DATA_FOLDER}result.xlsx")
    wb.close()


    con = sqlite3.connect(f"{DATA_FOLDER}result.db")                        # Создание базы данных, состоящей из двух таблиц.
    cur = con.cursor()
    # cur.execute("PRAGMA journal_mode=WAL")
    cur.execute("PRAGMA busy_timeout = 30000")
    cur.execute('''CREATE TABLE if not exists INVESTORS(
            investor_id integer,
            name_of_investor text,
            site text,
            info_card text,
            stage text text,
            check_size text,
            focus text,
            investment_geography text
    )''')

    cur.execute('''CREATE TABLE if not exists MANAGERS(
            manager_id integer primary key autoincrement,
            investor_id integer,
            manager_name text,
            role text,
            contacts text,
            foreign key (investor_id) references INVESTORS(investor_id) on delete cascade
    )''')
    con.commit()
    con.close()

def get_data_from_pages():
    options = webdriver.ChromeOptions()
    options.add_argument("user-agent=Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/106.0.0.0 Safari/537.36")
    options.add_argument("--disable-blink-features=AutomationControlled")
    options.headless = True
    
    try:
        driver = webdriver.Chrome(
            executable_path=PATH_TO_WEBDRIVER,
            options=options
         )
 
        count = 1                               #Счетчик инвесторов
        for page in range(0, 10, 10): 

            with open(f"{DATA_FOLDER}data/all_links - {page}.json") as file:
                all_links = json.load(file) 
            
            for name_fund in all_links: 

                try:
                   
                    driver.get(url=all_links[name_fund])
                    time.sleep(2)
                    wait = WebDriverWait(driver, 20).until(lambda x: x.find_element(By.ID, "ember7")) 
                    # wait = WebDriverWait(driver, 20)
                    # wait.until(visibility_of_element_located((By.ID, "ember36"))) 
                  

                    soup = BeautifulSoup(driver.page_source, 'lxml')
                   
                    print(f'{count}  {all_links[name_fund]} ... Готов!')

                    m_n =[]
                    m_r = []
                    m_l = []

                    link = soup.find(class_='mr-2 text-sm leading-tight text-orange-600 hover:underline').get('href').strip(' ')          
                    stage = soup.find(string=re.compile('Stage')).find_next('span').text.replace('\n', '').replace(' ', '')
                    check = soup.find(string=re.compile('Check size')).find_next('span').text.replace('\n', '').replace(' ', '')
                    focus = soup.find(string=re.compile('Focus')).find_next('span').string.replace(' ', '').replace(',', ', ').strip('\n ')
                    i_geo = soup.find(string=re.compile('Investment geography')).find_next('span').text.strip().replace('\n', '')
                    
                    m_name = soup.find_all(class_='font-serif text-base text-black truncate')

                    for i in m_name:
                        m_n.append(i.text.strip('\n '))
                    
                    m_role = soup.find_all(class_='text-xs text-gray-500 truncate')

                    for i in m_role:
                        m_r.append(i.text.strip('\n ') if i.text.strip('\n ') != '' else NO_INFO_STATUS)
                    
                    m_links = soup.find_all(class_='flex items-center px-2 py-3 text-sm bg-white border border-gray-300 max-w-sm')
                    
                    for i in m_links:
                        tmp = ''
                        for e in i.find_all('a'):
                            tmp += e.get('href') + ', '
                        m_l.append(tmp.strip('\n') if tmp.strip('\n') != '' else NO_INFO_STATUS)  

                    

                    add_row_to_table = [
                        count,
                        name_fund,
                        (link if link != '' else NO_INFO_STATUS),
                        all_links[name_fund],
                        (stage if stage != '' else NO_INFO_STATUS),
                        (check if check != '-' else NO_INFO_STATUS),
                        (focus if focus != '' else NO_INFO_STATUS),
                        (i_geo if i_geo != '' else NO_INFO_STATUS)
                        ]

                    con = sqlite3.connect(f"{DATA_FOLDER}result.db")                    
                    cur = con.cursor()
                    
                    cur.executemany('''insert into INVESTORS( investor_id,
                                                            name_of_investor,
                                                            site,
                                                            info_card,
                                                            stage,
                                                            check_size,
                                                            focus,
                                                            investment_geography) values(?, ?, ?, ?, ?, ?, ?, ?)''', (add_row_to_table,))
                    con.commit()

                    del add_row_to_table[0]                         # Удаление счетчика инвесторов

                    for num in range(len(m_n)):
                        mng = [count, m_n[num], m_r[num], m_l[num]]
                        cur.executemany('''insert into MANAGERS(
                                                            investor_id,
                                                            manager_name,
                                                            role,
                                                            contacts 
                                                            ) values(?, ?, ?, ?)''', (mng,) )
                        con.commit()
                       

                    count += 1                                     # Прибавляем счетчик инвесторов (investor_id)


                    for i in range(10):                             # Добавление всех сотрудников в общий список
                        try:
                            add_row_to_table.append(m_n[i])
                            add_row_to_table.append(m_r[i])
                            add_row_to_table.append(m_l[i])
                        except: 
                            add_row_to_table.append(NO_INFO_STATUS)
                            add_row_to_table.append(NO_INFO_STATUS)
                            add_row_to_table.append(NO_INFO_STATUS)             

                    with open(f"{DATA_FOLDER}result.csv", 'a') as file:          # Запись csv файла
                        writer = csv.writer(file)
                        writer.writerow(add_row_to_table)

                    wb = load_workbook(f"{DATA_FOLDER}result.xlsx")   # Добавление данных в таблицу excel 
                    ws= wb.active                                    
                    ws.append(add_row_to_table)                    
                    wb.save(f"{DATA_FOLDER}result.xlsx")
                     
                    
                
                         
                except Exception as ex:
                    print(ex)

       
           
    except Exception as ex:
        print(ex)

    finally:
        driver.close()
        driver.quit()
        con.close()    

def main():
    start = datetime.now()
    # get_all_links()
    create_result_headers()
    get_data_from_pages()
    print(f' Время выполнения ... {datetime.now() - start}')

if __name__ == '__main__':
    main()